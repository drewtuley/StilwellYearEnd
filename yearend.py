import configparser
from datetime import datetime

from openpyxl import load_workbook

mandatory_transaction_columns = {'Date', 'Payee', 'Amount'}

bank_credit_label = 'BANK - PAID IN'
bank_debit_label = 'BANK - PAID OUT'


def locate_text_in_worksheet(ws, text):
    f_col = 1
    row = 1
    f_row = None

    for col in ws.iter_cols(1, 20):
        for cell in col:
            if cell.value == text:
                f_row = row
                break
            else:
                row += 1
        if f_row is not None:
            break
        f_col += 1
        row = 1

    return format_cell_address(f_col+64, f_row)



def parse_cell_address(cell_addr):
    return ord(cell_addr[0]), int(cell_addr[1:])


def format_cell_address(col, row):
    return '{c}{r}'.format(c=chr(col), r=row)


def next_row(cell_addr, incr=1):
    curr_col, curr_row = parse_cell_address(cell_addr)
    return format_cell_address(curr_col, curr_row + incr)


def next_col(cell_addr, incr=1):
    curr_col, curr_row = parse_cell_address(cell_addr)
    return format_cell_address(curr_col + incr, curr_row)


def locate_next_free_cell_in_column(ws, cell_addr):
    while ws[cell_addr].value is not None:
        cell_addr = next_row(cell_addr)
    return cell_addr


def locate_next_free_row_in_range(ws, range_identifier):
    cell_addr = locate_text_in_worksheet(ws, range_identifier)
    return locate_next_free_cell_in_column(ws, cell_addr)


def determine_worksheet_name(dt, config):
    sheet_format = config.get('spreadsheet', 'sheet_name_format')

    return dt.strftime(sheet_format)


def parse_txn_date(s_date, config):
    txn_format = config.get('transaction', 'date_format')

    return datetime.strptime(s_date, txn_format)


def sheet_format_date(dt, config):
    sheet_date_format = config.get('spreadsheet', 'date_format')

    return dt.strftime(sheet_date_format)


def get_meta(header, config):
    meta = {}
    separator = config.get('transaction', 'separator')
    header_columns = header.split(separator)
    for col in mandatory_transaction_columns:
        try:
            idx = header_columns.index(col)
            meta[col] = idx
        except ValueError:
            print('Error: required header column {} does not exist'.format(col))
            exit(1)

    return meta


def category_mapping(text, config):
    try:
        # colons in the input string confuse the config parser - so remove them
        return config.get('category_mapping', text.replace(':', ''))
    except configparser.NoOptionError:
        return None


def process_txn(txn, meta, wb, config):
    separator = config.get('transaction', 'separator')
    fields = txn.split(separator)
    f_date = parse_txn_date(fields[meta['Date']], config)
    f_payee = fields[meta['Payee']]
    f_amount = float(fields[meta['Amount']])

    print('{d}_{p}_{a}'.format(d=f_date, p=f_payee, a=f_amount))
    ws_name = determine_worksheet_name(f_date, config)
    ws = wb[ws_name]
    ws.protection.disable()

    if f_amount > -1:
        # Credit
        cell_addr = locate_next_free_row_in_range(ws, bank_credit_label)
        ws[cell_addr].value = sheet_format_date(f_date, config)
        cell_addr = next_col(cell_addr)

        ws[cell_addr].value = f_payee
        cell_addr = next_col(cell_addr, incr=2)
        ws[cell_addr].value = f_amount
        cell_addr = next_col(cell_addr)
        category = category_mapping(f_payee, config)
        if category is not None:
            ws[cell_addr].value = category

    else:
        # debit
        cell_addr = locate_next_free_row_in_range(ws, bank_debit_label)
        ws[cell_addr].value = sheet_format_date(f_date, config)
        cell_addr = next_col(cell_addr)
        ws[cell_addr].value = f_payee
        cell_addr = next_col(cell_addr)
        category = category_mapping(f_payee, config)
        if category is not None:
            ws[cell_addr].value = category

        cell_addr = next_col(cell_addr)
        ws[cell_addr].value = f_amount * -1


if __name__ == "__main__":
    config = configparser.ConfigParser()
    config.read('yearend.props')

    wb = load_workbook('Accounts Spreadsheet (FRS) -Elref Ltd -July 2018.xlsx')
    txns = config.get('transaction', 'filename')
    with open(txns) as fd:
        meta = None
        for rtxn in fd:
            txn = rtxn.strip()
            if meta is None:
                meta = get_meta(txn, config)
                if meta is None:
                    print('Unable to parse meta')
                    exit(1)
            else:
                process_txn(txn, meta, wb, config)
    sav_file = config.get('spreadsheet', 'save_filename')
    wb.save(sav_file)
